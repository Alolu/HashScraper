# -*- coding: utf-8 -*-
import bs4 as bs
import requests as r
import time,re,json,sys,random,datetime
from openpyxl import load_workbook
import sqlalchemy as db

pattern = re.compile('window._sharedData = (.*?});') #Pattern de la variable contenant les infos pertinente dans les pages instagram
instagram_tag_url = "https://www.instagram.com/explore/tags/" #Url de l'exploration d'hashtag instagram
hashtag_generator_url = "https://www.all-hashtag.com/library/contents/ajax_generator.php" #Url du generateur de hashtag similaire
tag_to_test = sys.argv[1] #Le premier argument est le hashtag de base sur le quel on souhait baser la generation de hashtag similaire
titleOffset = 4

engine = db.create_engine('mysql+pymysql://root:@localhost/hashtag_cache')
metadata = db.MetaData()
metadata.reflect(bind=engine)

groups_table = db.Table('groups', metadata, autoload=True, autoload_with=engine)
users_table = db.Table('clients', metadata, autoload=True, autoload_with=engine)
hashtags_table = db.Table('hashtags', metadata, autoload=True, autoload_with=engine)
used_table = db.Table('used', metadata, autoload=True, autoload_with=engine)

number_of_tags = 30
hashtag_list = []
expired_tags = set()

class Hashtag:
    def __init__(self,holder):
        self.hashtag = holder["hashtag"]
        self.link = holder["link"]
        self.total_posts = holder["total_posts"]
        self.avg_likes = holder["avg_likes"]
        self.avg_comms = holder["avg_comms"]
        self.min_likes = holder["min_likes"]
        self.max_likes = holder["max_likes"]
        self.min_comms = holder["min_comms"]
        self.max_comms = holder["max_comms"]
        self.expire = datetime.datetime.now() + datetime.timedelta(days=3)

    def set_id_group(self,group_id):
        self.id_group = group_id
    def update(self):
        update(hashtags_table,vars(self),hashtags_table.c.hashtag == self.hashtag)
        print(f"#Database->Expired hashtag {self.hashtag} updated in database")
    def save(self):
        push(hashtags_table,vars(self))
        print(f"#Database->Hashtag {self.hashtag} inserted in database")

    def __str__(self):
        return f"""
        Hashtag form : { self.hashtag }
        Url address : { self.link }
        Total posts count : { self.total_posts }
        Average likes : { self.avg_likes }
        Minimum likes : { self.min_likes }
        Maximum likes : { self.max_likes }
        Average comments : { self.avg_comms }
        Minimum comments : { self.min_comms }
        Maximum comments : { self.max_comms }
        """

class HashtagDownloader:
    """Une classe permettant de recuperer toute les info sur un hashtag, requiert un hashtag complet a l'initialisation"""
    def __init__(self,hashtag):
        self.hashtag = hashtag
        self.name = hashtag[1:]
        self.link = instagram_tag_url + self.name
        self.shared_data = self.__download_hashtag()
        self.top_posts = self.__extract_tops()
        self.avg_likes, self.avg_comms, self.max_likes, self.min_likes, self.max_comms, self.min_comms = self.__extract_likes_comments()
        self.total_posts = self.__extract_total_posts()

    def output(self):
        return {
            "hashtag":self.hashtag,
            "link":self.link,
            "total_posts":self.total_posts,
            "avg_likes":self.avg_likes,
            "avg_comms":self.avg_comms,
            "min_likes":self.min_likes,
            "max_likes":self.max_likes,
            "min_comms":self.min_comms,
            "max_comms":self.max_comms,
        }

    def __download_hashtag(self):
        """Fonction privé récupérant la data depuis la page instagram du hashtag"""
        print(f"HashtagInfo->Fetching data for {self.hashtag}.")
        src = r.get(self.link)
        soup = bs.BeautifulSoup(src.content.decode('utf-8','ignore'),'html.parser')
        scripts  = soup.find_all("script")
        for script in scripts:
            if(pattern.match(str(script.string))):
                data = pattern.match(script.string)
                print(f"HashtagInfo->Data fetched for {self.hashtag}.")
                return json.loads(data.groups()[0])
    def __extract_tops(self):
        """Fonction permettant d'extraire les 9 meilleurs posts contenu dans la data téléchargée"""
        return (self.shared_data
            .get("entry_data")
            .get("TagPage")[0]
            .get("graphql")
            .get("hashtag")
            .get("edge_hashtag_to_top_posts")
            .get("edges"))
    def __extract_likes_comments(self):
        """Fonction permettant d'extraire la moyenne, le maximum et le minimum
        des likes et commentaires des 9 meilleurs posts contenu dans la data téléchargée"""
        #A refactorer en 3 fonctions differentes...
        total_likes = 0
        total_comms = 0
        max_like = float('-inf')
        min_like = float('inf')
        max_comm = float('-inf')
        min_comm = float('inf')
        for post in self.top_posts:
            post = post.get("node")
            likes = post.get("edge_liked_by").get("count")
            comms = post.get("edge_media_to_comment").get("count")
            total_likes += likes
            total_comms += comms
            if likes < min_like:
                min_like = likes
            if likes > max_like:
                max_like = likes

            if comms < min_comm:
                min_comm = comms
            if comms > max_comm:
                max_comm = comms
            
        return int(total_likes/len(self.top_posts)), int(total_comms/len(self.top_posts)), max_like, min_like, max_comm, min_comm
    def __extract_total_posts(self):
        """Fonction permettant d'extraire le nombre de post total contenu depuis data téléchargée"""
        return (self.shared_data
            .get("entry_data")
            .get("TagPage")[0]
            .get("graphql")
            .get("hashtag")
            .get("edge_hashtag_to_media")
            .get("count")
        )
    def __str__(self):
        return f"""
        Name : { self.name }
        Hashtag form : { self.hashtag }
        Url address : { self.link }
        Total posts count : { self.total_posts }
        Average likes : { self.avg_likes }
        Minimum likes : { self.min_likes }
        Maximum likes : { self.max_likes }
        Average comments : { self.avg_comms }
        Minimum comments : { self.min_comms }
        Maximum comments : { self.max_comms }
        """

def to_dict(keys,result):
    dict_list = []
    for val in result:
        dict_object = {}
        for li in val.items():
            dict_object[li[0]] = li[1]
        dict_list.append(dict_object)
    return dict_list
def get_where(table,condition):
    connection = engine.connect()
    query = db.select([table]).where(condition)
    result_proxy = connection.execute(query)
    return to_dict(table.columns.keys(),result_proxy.fetchall())
def get_all(table):
    connection = engine.connect()
    query = db.select([table])
    result_proxy = connection.execute(query)
def push(table,data):
    connection = engine.connect()
    query = db.insert(table).values(data)
    result_proxy = connection.execute(query)
def separator(amount=30):
    print("\n"+("-"*amount)+ "\n")
def update(table,value,condition):
    connection = engine.connect()
    query = table.update().values(value).where(condition)
    result_proxy = connection.execute(query)
def get_related_hashtags(tag):
    """Fonction permettant de generer des hashtags similaire a celui passé en argument en scrapant all-hashtag, prends en parametre un hashtag incomplet"""
    src = r.post(hashtag_generator_url,{"keyword":tag,"filter":"random"})
    soup = bs.BeautifulSoup(src.content.decode("utf-8","ignore"),"html.parser")
    tags = soup.find('div',{"class":"copy-hashtags"}).get_text().lower()
    print("#HashtagGatherer->Related hashtags fetched")
    return set(tags.split())
def download_hashtags_content(group_id,hash_list,update=False,minTime = 5, maxTime = 12.5):
    """Fonction qui analyse entierement une liste de hashtag et attends 10 secondes entre chaque 
    recupération d'info d'hashtag, prends comme argument optionnel un float, qui est par default a 10"""
    for i,hashtag in enumerate(hash_list,0):
        hsd = HashtagDownloader(hashtag)
        hs = Hashtag(hsd.output())
        hs.set_id_group(group_id)
        hs.update() if update else hs.save()
        hashtag_list.append(hs)
        #write_in_line(i,analyzed_hashtag)
        waitingTime = random.uniform(minTime,maxTime)
        print(f"#Antiblocker->Waiting {waitingTime} for next Hashtag...\n")
        time.sleep(waitingTime)
def write_in_cell(cell,value):
    ws1.cell(cell.row,cell.column,value)
def write_in_line(line,hashtag):
    print(f"Workbook->Writing line for {hashtag.hashtag}")
    line += titleOffset
    cells = ws1.iter_cols(2,10,line,line)
    for value in hashtag.output():
        write_in_cell(next(cells)[0],value)
    print(f"Workbook->Line written for {hashtag.hashtag}")
def add_tags_from_database(group):
    global number_of_tags

    print(f"#Database->Hashtag {tag_to_test} found in records")
    tags_from_base = get_where(hashtags_table,hashtags_table.c.id_group == group['id'])
    for tag in tags_from_base:
        if(tag['expire'] < datetime.datetime.now()):
            expired_tags.add(tag['hashtag'])
            print(f"#Alert->Hashtag { tag.get('hashtag') } is expired")
        else:
            hs = Hashtag(tag)
            hs.set_id_group(tag["id_group"])
            hashtag_list.append(hs)
            number_of_tags -= 1
    print(f"#HashtagManager->Remaining hashtag to fetch: {number_of_tags}")
def add_group_to_database():
    print(f"#Database->Hastag {tag_to_test} not found in records")
    push(groups_table,{"name":tag_to_test})
    print(f"#Database->Hashtag {tag_to_test} saved")
    #Très très moche,, a revamp dans le futur
    return get_where(groups_table,groups_table.c.name == tag_to_test)


def main():
    #wb = load_workbook("Template.xlsx")
    #ws1 = wb.worksheets[0]
    separator()
    print(f"Hashtag group: {tag_to_test}")
    print(f"Amount: {number_of_tags}")
    separator()
    print(f"#Database->Checking if group {tag_to_test} is registered...")
    #A AMELIOREEEEEEEER
    group_data = get_where(groups_table,groups_table.c.name == tag_to_test)
    if group_data:
        add_tags_from_database(group_data[0])  
    else:
        group_data = add_group_to_database()
    group = group_data[0]
    separator()
    related_hashtags = get_related_hashtags(tag_to_test) - expired_tags
    print("#HashtagGatherer->Doubles trimmed")
    separator()
    print("#HashtagManager->Updating expired hashtags\n")
    download_hashtags_content(group['id'],expired_tags,True)
    separator()
    print("#HashtagManager->Downloading remaining hashtags\n")
    download_hashtags_content(group['id'],related_hashtags)
    #hashtag_list = get_related_hashtags(tag_to_test)
    #get_hashtags_info()
    #wb.save("Template.xlsx")

main()

